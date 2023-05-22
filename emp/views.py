from django.shortcuts import render
from openpyxl import load_workbook
from .forms import RegEmpForm

def regEmp(request):
    m=''
    form = RegEmpForm()

    if request.method == 'POST':
        form = RegEmpForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            wb = load_workbook(r'C:\Users\ultra1\Desktop\learning dj\kareem dj\srcKareemDj\names.xlsx')
            ws = wb['emp']

            i = str(int(ws['M1'].value) + 1)

            ws['A' + i] = cd['name']
            ws['D'+ i] = cd['salary']
            ws['G' + i] = cd['emp_email']
            ws['J' + i] = cd['address']

            ws['M1'] = i
            wb.save(r'C:\Users\ultra1\Desktop\learning dj\kareem dj\srcKareemDj\names.xlsx')
            output = 'تم الحفظ بنجاح'
            form = RegEmpForm()

    return render(request, 'emp/newemp.html', {'form': form, 'output': m})

