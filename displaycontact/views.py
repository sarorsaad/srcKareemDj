from django.shortcuts import render
from openpyxl import load_workbook

def disp(request):
    wb = load_workbook(r'C:\Users\ultra1\Desktop\learning dj\kareem dj\srcKareemDj\names.xlsx')
    ws = wb['emp']
    noofrows = int(ws['M1'].value)
    items = []
    for i in range(2, noofrows + 1):
        item = []
        item.append(ws['A' + str(i)].value)
        item.append(ws['D' + str(i)].value)
        item.append(ws['G' + str(i)].value)
        item.append(ws['J' + str(i)].value)
        items.append(item)
    return render(request, 'displaycontact/displaydata.html', {'items': items})
